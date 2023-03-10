import requests
from bs4 import BeautifulSoup
import datetime 
import openpyxl
import os



now = datetime.datetime.now()
today = now.strftime('%Y-%m-%d')


new_file = r'C:\Users\PC\Desktop\xo0ol_files\crawling\lotte_crawling({}).xlsx'.format(today)
wb = openpyxl.Workbook()
ws = wb.create_sheet('lotte')

uniq = 1
while os.path.exists(new_file):
    new_file = r'C:\Users\PC\Desktop\xo0ol_files\crawling\lotte_crawling({})({}).xlsx'.format(today, uniq)
    uniq += 1



ref_file = r'C:\Users\PC\Desktop\xo0ol_files\crawling\(Ref)hmp_mall_crawling.xlsx'
wb = openpyxl.load_workbook(ref_file)
ws = wb['hmp']
print(ws)

j = 1
hmp_url = []
for i in ws.rows:
    hmp_url.append(ws[f'a{j}'].value)
    j = j + 1

url_13code = []
for x in hmp_url:
    x = str(x)
    url_13code.append(x.rjust(9,"0"))

print(url_13code)

j = 1
for i in url_13code:
    try:
        respon = requests.get(f"https://mfront.homeplus.co.kr/item?itemNo={i}&storeType=HYPER")
        html = respon.text
        soup = BeautifulSoup(html, 'html.parser')
        title = soup.select_one('title').text
        title = title.replace(' | 홈플러스','')
        
    except:
        title = '홈플러스 판매종료'

        
    print(title)
    ws[f'a{j}'] = title
    j = j+1



wb.save(new_file)
