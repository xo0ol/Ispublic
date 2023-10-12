import requests
from bs4 import BeautifulSoup
import os
import openpyxl
import math


import time # time.sleep() 을 위한 패키지
import timeit # 시간을 숫자단위로 측정. 시작시간-종료시간으로 작업시간을 계산
from datetime import datetime # datetime.now() 을 위한 패키지
from datetime import timedelta # 시간끼리의 연산을 위한 패키지


# running time check
start_time = timeit.default_timer()




## 이곳에서 ref file과 new file의 경로와 이름을 설정하시오. ##

# 1.새로 만들 파일을 저장할 주소와 이름을 설정하시오.
new_file_name = '\koreanet_crawling'


new_file_adress = r'C:\Users\xo0ol\OneDrive\바탕 화면\xoyoung\crawling\koreanet'
# new_file_adress = r'C:\Users\user\Desktop\소영파일\koreanet' # 궁동집
# new_file_adress = r'C:\Users\PC\OneDrive\바탕 화면\xoyoung\crawling\koreanet' # 궁동집안방


# 2. 바코드 정보를 가져올 파일의 주소와 시트명을 설정하시오.
ref_file = r'C:\Users\xo0ol\OneDrive\바탕 화면\xoyoung\crawling\koreanet.txt'
# ref_file = r'C:\Users\user\Desktop\소영파일\koreanet.txt' # 궁동집
# ref_file = r'C:\Users\PC\OneDrive\바탕 화면\xoyoung\crawling\koreanet.txt' # 궁동집안방

############################################################





# txt 바코드 파일에서 바코드 데이터 리스트로 가져오기
file = open(ref_file, 'r')
read = file.readlines()

barcode = [x.strip('\n') for x in read]
print(barcode)
count_bar = len(barcode)





# excel 바코드 파일에서 바코드 데이터 리스트로 가져오기
# wb = openpyxl.load_workbook(ref_file)
# ws = wb['barcode']

# j = 1
# barcode = []
# for i in ws.rows:
#     barcode.append(ws[f'a{j}'].value)
#     j = j+1

# k = 1
# while k <= len(barcode):
#     if None in barcode:
#         s = barcode.index(None)
#         barcode[s] = ""
#     k = k + 1







# 워크시트 생성
wb_2 = openpyxl.Workbook()
ws_2 = wb_2.active
idx = 1




# 웹 크롤링 시작
for x in barcode:
    try:
        url = 'http://www.koreannet.or.kr/home/hpisSrchGtin.gs1?gtin=' + str(x)
        respon = requests.get(url)
        html = respon.text
        soup = BeautifulSoup(html, 'html.parser')

        title = soup.select_one('.productTit').text.strip()

        # 중량 찾기
        ss = soup.find('dd', class_='productStd clfix')
        find_inf = ss.find_all('th')
        find_weight = ss.find_all('td')
        

        inf = []
        weight = []
        weight_2 = []


        for i in find_inf:
            inf.append(i.text)



        # inf에 순중량이 있다면 중량찾기 실행됨.

        for i in find_weight:
            weight.append(i.text)

        inf_num = inf.index("순중량")

        weight = weight[inf_num]



        
        # inf에 용량이 있다면 중량찾기 실행됨.

        for i in find_weight:
            weight_2.append(i.text)

        inf_num = inf.index("용량")

        weight_2 = weight_2[inf_num]



        # 데이터 저장
        today = (datetime.now()).strftime('%Y-%m-%d %H:%M:%S')

        if len(weight_2) <= 1 :
            ws_2[f'a{idx}'] = f'koreannet_search/{str(title)} | {weight} | {today}'
            print(f'({idx}/{count_bar}) koreannet_search/{str(title)} | {weight} | {today}')

        elif weight == ' G':
            ws_2[f'a{idx}'] = f'koreannet_search/{str(title)} | {weight_2} | {today}'
            print(f'({idx}/{count_bar}) koreannet_search/{str(title)} | {weight_2} | {today}')

        else:
            ws_2[f'a{idx}'] = f'koreannet_search/{str(title)} | {weight} | {weight_2} | {today}'
            print(f'({idx}/{count_bar}) koreannet_search/{str(title)} | {weight} | {weight_2} | {today}')
            
           
    except:
        today = (datetime.now()).strftime('%Y-%m-%d %H:%M:%S')
        ws_2[f'a{idx}'] = f"no product({str(x)}) {today}"
        # ws_2[f'a{idx}'] = "koreannet_search/" + str(x).ljust(13," ") + "no product"
        # print('('+str(idx)+'/'+str(count_bar)) + f"no product({str(x)})"
        print(f'({str(idx)}/{str(count_bar)}) no product({str(x)}) {today}')
        
    idx +=1
    time.sleep(0.5)




# new file 저장하기.
now = datetime.now()
today = now.strftime('%Y-%m-%d')
new_file = new_file_adress  + new_file_name + '({}).xlsx'.format(today)


uniq = 1
while os.path.exists(new_file):
    new_file = new_file_adress  + new_file_name + '({})({}).xlsx'.format(today, uniq)
    uniq += 1

wb_2.save(new_file)






# 작업 종료 알림
end_now_str = (datetime.now()).strftime('%Y-%m-%d %H:%M:%S')
finished_time = timeit.default_timer()
running_time = math.trunc(finished_time - start_time)


if running_time % 60 == running_time:
    print(running_time)
    x_time = f'00:{str(running_time).rjust(2,"0")}'
else:
    x1 = math.trunc(running_time / 60)
    x2 = running_time & 60
    x_time = f'{str(x1).rjust(2,"0")}:{str(x2).rjust(2,"0")}'

print(f'『 {(x_time)} 소요되었습니다. 』')
print(f"『 {end_now_str} 작업을 종료합니다. 』")