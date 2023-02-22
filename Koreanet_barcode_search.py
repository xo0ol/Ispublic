import requests
from bs4 import BeautifulSoup
import datetime
import os
import openpyxl
import timeit
import time

start_time = timeit.default_timer()

# 바코드 파일
ref_file = r'C:\Users\sjeon\OneDrive\바탕 화면\xoyoung\crawling\(Ref)hmp_search_data.xlsx'


# 생성 파일
now = datetime.datetime.now()
today = now.strftime('%Y-%m-%d')
new_file = r'C:\Users\sjeon\OneDrive\바탕 화면\xoyoung\crawling\hmp_search_crawling\koreanet_crawling({}).xlsx'.format(today) 

uniq = 1
while os.path.exists(new_file):
    new_file = r'C:\Users\sjeon\OneDrive\바탕 화면\xoyoung\crawling\hmp_search_crawling\koreanet_crawling({})({}).xlsx'.format(today, uniq)
    uniq += 1


# 바코드 파일에서 바코드 데이터 리스트로 가져오기
wb = openpyxl.load_workbook(ref_file)
ws = wb['barcode']


j = 1
barcode = []
for i in ws.rows:
    barcode.append(ws[f'a{j}'].value)
    j = j+1



k = 1
while k <= len(barcode):
    if None in barcode:
        s = barcode.index(None)
        barcode[s] = ""
    k = k + 1
print(barcode)


count_bar = len(barcode)
print(count_bar)




# 워크시트 생성
wb_2 = openpyxl.Workbook()
ws_2 = wb_2.active
idx = 1


for x in barcode:
    try:
        url = 'http://www.koreannet.or.kr/home/hpisSrchGtin.gs1?gtin=' + str(x)
        respon = requests.get(url)
        html = respon.text
        soup = BeautifulSoup(html, 'html.parser')

        title = soup.select_one('.productTit').text.strip()

        # 중량 찾기
        ss = soup.find('dd', class_='productStd clfix')
        find_inf = ss.find_all('th')
        find_weight = ss.find_all('td')
        

        inf = []
        weight = []
        weight_2 = []


        for i in find_inf:
            inf.append(i.text)



        # inf에 순중량이 있다면 중량찾기 실행됨.

        for i in find_weight:
            weight.append(i.text)

        inf_num = inf.index("순중량")

        weight = weight[inf_num]



        
        # inf에 용량이 있다면 중량찾기 실행됨.

        for i in find_weight:
            weight_2.append(i.text)

        inf_num = inf.index("용량")

        weight_2 = weight_2[inf_num]



        # 데이터 저장
        if len(weight_2) <= 1 :
            ws_2[f'a{idx}'] = "koreannet_search/" + str(title) + " | " + weight
            print(f'({idx}/{count_bar}) koreannet_search/'+str(title) + " | " + weight)

        elif weight == ' G':
            ws_2[f'a{idx}'] = "koreannet_search/" + str(title) + " | " + weight_2
            print(f'({idx}/{count_bar}) koreannet_search/'+str(title) + " | " + weight_2) 

        else:
            ws_2[f'a{idx}'] = "koreannet_search/" + str(title) + " | " + weight + " | " + weight_2
            print('('+str(idx)+'/'+str(count_bar)+') koreannet_search/'+str(title) + " | " + weight + " | "+ weight_2)
           
    except:

        ws_2[f'a{idx}'] = "koreannet_search/" + str(x).ljust(13," ") + "      no product"
        print('('+str(idx)+'/'+str(count_bar)+') koreannet_search/'+ str(x).ljust(13," ") + "      no product")
        
    idx +=1
    time.sleep(0.5)



wb_2.save(new_file)

terminate_time = timeit.default_timer()
print("%f초 걸렸습니다." % (terminate_time - start_time))
