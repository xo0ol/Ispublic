import openpyxl
import pyautogui
import time
import datetime
import requests
from bs4 import BeautifulSoup
import os




headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36'}


ref_file = r'C:\Users\PC\Desktop\xo0ol_files\crawling\(Ref)hmp_mall_crawling.xlsx'
wb = openpyxl.load_workbook(ref_file)
ws = wb['emart']



j = 1
e_url = []
for i in ws.rows:
    e_url.append(ws[f'a{j}'].value)
    j = j + 1



# e_url = list(filter(None, e_url))
# print(e_url)

k = 1
while k <= len(e_url):
    if None in e_url:
        s = e_url.index(None)
        e_url[s] = "https://emart.ssg.com/"
    k = k + 1
print(e_url)
count = len(e_url)
print(count)

# new file
now = datetime.datetime.now()
today = now.strftime('%Y-%m-%d')
new_file = r'C:\Users\PC\Desktop\xo0ol_files\crawling\Emart_crawling\emart_crawling({}).xlsx'.format(today)

uniq = 1
while os.path.exists(new_file):
    new_file = r'C:\Users\PC\Desktop\xo0ol_files\crawling\Emart_crawling\emart_crawling({})({}).xlsx'.format(today, uniq)
    uniq += 1



wb_new = openpyxl.Workbook()
ws_new = wb_new.active
ws_new['a1'] = 'title'
ws_new['b1'] = 'new_price'
ws_new['c1'] = 'price'
ws_new['d1'] = 'delivery'
ws_new['e1'] = 'out_of_stock'
ws_new['f1'] = 'promotion'



rows = 2
idx = 1
for i in e_url:

    requests.get(i, headers=headers)
    respon = requests.get(i, headers=headers)
    html = respon.text
    soup = BeautifulSoup(html, 'html.parser')



    try:
        title = soup.select_one("div.cdtl_prd_info > h2").text.strip()
    except:
        ws_new[f'a{rows}'] = "X"
        ws_new[f'b{rows}'] = ""
        ws_new[f'c{rows}'] = ""
        ws_new[f'd{rows}'] = ""
        ws_new[f'e{rows}'] = ""
        ws_new[f'f{rows}'] = ""
        print('('+str(count)+'/'+str(idx)+') 완료')
    else:     
        try:
            new_price = soup.select_one(".cdtl_new_price").text.strip()
        except:
            new_price = "오류"
        
        try:
            price = soup.select_one(".ssg_price").text
        except:
            price = "오류"

        s = soup.select("span > span")
        ss = []
        for i in s:
            ss.append(i.text)

        if '품절' in ss:
            out_of_stock = '품절'
        elif '일시품절' in ss:
            out_of_stock = '일시품절'
        else:
            out_of_stock = 0

        try:
            if soup.select_one("dl.cdtl_dl.cdtl_delivery_fee > dt").text is not None:
                delv = '택배배송'
        except:
            try:
                if soup.select_one(".cdtl_delv_type").text is not None:
                    delv = 'ssg delivery'
            except:
                delv = ""

        try:
            promotion = soup.find("div",{"class":"txt"}).text.strip()
        except:
            promotion = ""

        print(title.strip())
        print(new_price.strip())
        print(price.strip())
        print(str(out_of_stock).strip())
        print(delv.strip())
        print(promotion.strip())
        print('('+str(count)+'/'+str(idx)+') 완료')
        ws_new[f'a{rows}'] = title
        ws_new[f'b{rows}'] = new_price
        ws_new[f'c{rows}'] = price
        ws_new[f'd{rows}'] = delv
        ws_new[f'e{rows}'] = out_of_stock
        ws_new[f'f{rows}'] = promotion
    rows += 1
    idx += 1

    time.sleep(5)

wb_new.save(new_file)

