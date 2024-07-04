import time
import requests
from bs4 import BeautifulSoup
import pyautogui
import pyperclip
import openpyxl
import datetime
import os

s = 123

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

# 크롬 드라이버 자동 업데이트
from webdriver_manager.chrome import ChromeDriverManager

# 브라우저 꺼짐 방지
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
service = Service(executable_path=ChromeDriverManager().install())

# 불필요한 에러 메세지 없애기
chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
chrome_options.add_argument('--headless')


# 파일명 정의
ref_file = r'C:\Users\PC\Desktop\xo0ol_files\crawling\(Ref)hmp_search_crawling.xlsx' # 울집컴
# ref_file = r'C:\Users\sjeon\OneDrive\바탕 화면\xoyoung\crawling\(Ref)hmp_search_data.xlsx' #젠북

now = datetime.datetime.now()
today = now.strftime('%Y-%m-%d')
new_file = r'C:\Users\PC\Desktop\xo0ol_files\crawling\hmp_search_crawling\search_crawling({}).xlsx'.format(today) #울집컴
# new_file = r'C:\Users\sjeon\OneDrive\바탕 화면\xoyoung\crawling\hmp_search_crawling\hmp_search({}).xlsx'.format(today) # 젠북

uniq = 1
while os.path.exists(new_file):
    new_file = r'C:\Users\PC\Desktop\xo0ol_files\crawling\hmp_search_crawling\search_crawling({})({}).xlsx'.format(today, uniq)
    uniq += 1



wb = openpyxl.load_workbook(ref_file)
ws = wb['barcode']


j = 1
barcode = []
for i in ws.rows:
    barcode.append(ws[f'a{j}'].value)
    j = j+1


k = 1
while k <= len(barcode):
    if None in barcode:
        s = barcode.index(None)
        barcode[s] = ""
    k = k + 1

print(barcode)
count_bar = len(barcode)
print(count_bar)

wb_2 = openpyxl.Workbook()
ws_2 = wb_2.active
con = 1



# 주소 이동
browser = webdriver.Chrome(service=service, options=chrome_options)
url = 'https://www.consumer.go.kr/user/ftc/consumer/goodsinfo/57/selectGoodsInfoList.do'
browser.get(url)

browser.find_element(By.XPATH,'//*[@id="column"]/option[text()="유통표준코드"]').click()

for b in barcode:
    try:
        search = browser.find_element(By.ID,"search")
        search.clear()
        search.send_keys(b)
        search_button = browser.find_element(By.CLASS_NAME, 'btn_submit')
        search_button.click()
        time.sleep(2)

        titles = browser.find_element(By.CSS_SELECTOR, "#m_page_top > div.goodsViewDescBox > div > h3").text
        find_inf = browser.find_elements(By.CSS_SELECTOR,"div > dl > dt > span")
        find_weight = browser.find_elements(By.CSS_SELECTOR, "dl > dd")

        inf_data = []
        weight_data = []

        for i in find_inf:
            inf_data.append(i.text)

        if "중량" in inf_data:
            for i in find_weight:
                weight_data.append(i.text)

            inf_num = inf_data.index("중량")

            weight = weight_data[inf_num]
        else:
            weight = "No product"
        
        
        if weight == "No product":
            print(f"({count_bar}/{con}) 완료 koreannet/search/" + str(b) + " : " + titles)
            ws_2[f'a{con}'] = "koreannet_search/" + str(b) + " : " + titles
        else:
            print(f"({count_bar}/{con}) 완료 koreannet/search/" + str(b) + " : " + titles + " | " + weight)
            ws_2[f'a{con}'] = "koreannet_search/" + str(b) + " : " + titles + " | " + weight



    except:
        print(f"({count_bar}/{con}) 완료 koreannet_search/" + str(b) + " : No product")
        ws_2[f'a{con}'] = "koreannet_search/" + str(b) + " : No product"
        next
    
    con = con + 1
    browser.back()
    time.sleep(1)

browser.quit()      
wb_2.save(new_file)
